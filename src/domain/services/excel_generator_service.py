"""
Servicio para generación de archivos Excel

Este servicio crea archivos Excel profesionales con múltiples hojas,
formato profesional, filtros automáticos y freeze de filas.
"""
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelGeneratorService:
    """Servicio para generar archivos Excel profesionales"""
    
    def __init__(self):
        self.header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generar_excel_control(
        self, 
        control_nombre: str,
        consultas_resultados: List[Dict[str, Any]],
        referente_path: str,
        fecha_ejecucion: datetime = None
    ) -> str:
        """
        Genera un archivo Excel con los resultados de las consultas de un control
        
        Args:
            control_nombre: Nombre del control ejecutado
            consultas_resultados: Lista de diccionarios con los resultados de cada consulta
                Formato: [{'nombre': 'consulta1', 'datos': [dict], 'columnas': [str]}, ...]
            referente_path: Ruta donde guardar el archivo
            fecha_ejecucion: Fecha de ejecución (por defecto now)
            
        Returns:
            str: Ruta completa del archivo generado
        """
        if fecha_ejecucion is None:
            fecha_ejecucion = datetime.now()
        
        # Crear nombre de archivo
        timestamp = fecha_ejecucion.strftime("%Y%m%d_%H%M%S")
        # Limpiar nombre del control para usar en archivo
        control_limpio = self._limpiar_nombre_archivo(control_nombre)
        filename = f"{control_limpio}_{timestamp}.xlsx"
        filepath = os.path.join(referente_path, filename)
        
        # Crear directorio si no existe
        os.makedirs(referente_path, exist_ok=True)
        
        # Crear libro de Excel
        workbook = Workbook()
        
        # Eliminar hoja por defecto
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Crear hoja de resumen
        self._crear_hoja_resumen(workbook, control_nombre, consultas_resultados, fecha_ejecucion)
        
        # Crear hoja para cada consulta con datos
        for i, consulta_resultado in enumerate(consultas_resultados):
            if consulta_resultado.get('datos') and len(consulta_resultado['datos']) > 0:
                self._crear_hoja_consulta(workbook, consulta_resultado, i + 1)
        
        # Guardar archivo
        workbook.save(filepath)
        
        print(f"DEBUG - Archivo Excel generado: {filepath}")
        return filepath
    
    def _crear_hoja_resumen(
        self, 
        workbook: Workbook, 
        control_nombre: str, 
        consultas_resultados: List[Dict[str, Any]],
        fecha_ejecucion: datetime
    ):
        """Crea la hoja de resumen del control"""
        ws = workbook.create_sheet("Resumen", 0)
        
        # Título
        ws['A1'] = f"Reporte de Control: {control_nombre}"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
        ws.merge_cells('A1:C1')
        
        # Información general
        row = 3
        ws[f'A{row}'] = "Fecha de Ejecución:"
        ws[f'B{row}'] = fecha_ejecucion.strftime("%d/%m/%Y %H:%M:%S")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total de Consultas:"
        ws[f'B{row}'] = len(consultas_resultados)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Consultas con Datos:"
        consultas_con_datos = sum(1 for c in consultas_resultados if c.get('datos') and len(c['datos']) > 0)
        ws[f'B{row}'] = consultas_con_datos
        ws[f'A{row}'].font = Font(bold=True)
        
        # Tabla de consultas
        row += 3
        ws[f'A{row}'] = "Consulta"
        ws[f'B{row}'] = "Filas"
        ws[f'C{row}'] = "Estado"
        
        # Formatear encabezados
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Datos de consultas
        for consulta in consultas_resultados:
            row += 1
            ws[f'A{row}'] = consulta.get('nombre', 'Sin nombre')
            ws[f'B{row}'] = len(consulta.get('datos', []))
            ws[f'C{row}'] = "Con datos" if consulta.get('datos') else "Sin datos"
            
            # Bordes para datos
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = self.border
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _crear_hoja_consulta(
        self, 
        workbook: Workbook, 
        consulta_resultado: Dict[str, Any],
        numero_consulta: int
    ):
        """Crea una hoja para una consulta específica"""
        nombre_consulta = consulta_resultado.get('nombre', f'Consulta_{numero_consulta}')
        # Limpiar nombre para que sea válido como nombre de hoja
        nombre_hoja = self._limpiar_nombre_hoja(nombre_consulta)
        
        ws = workbook.create_sheet(nombre_hoja)
        
        datos = consulta_resultado.get('datos', [])
        if not datos:
            ws['A1'] = "No hay datos para mostrar"
            return
        
        # Obtener columnas
        columnas = list(datos[0].keys()) if datos else []
        
        # Crear encabezados
        for col_idx, columna in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=columna)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Agregar datos
        for row_idx, fila in enumerate(datos, 2):
            for col_idx, columna in enumerate(columnas, 1):
                valor = fila.get(columna, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = self.border
        
        # Freeze primera fila
        ws.freeze_panes = 'A2'
        
        # Crear tabla con autofiltro
        if len(datos) > 0:
            table_range = f"A1:{self._get_column_letter(len(columnas))}{len(datos) + 1}"
            # Limpiar nombre de tabla (no puede tener espacios ni caracteres especiales)
            table_name = self._limpiar_nombre_tabla(f"Tabla_{nombre_hoja}")
            table = Table(displayName=table_name, ref=table_range)
            
            # Estilo de tabla
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            ws.add_table(table)
        
        # Ajustar ancho de columnas
        for col_idx, columna in enumerate(columnas, 1):
            # Calcular ancho basado en el contenido
            max_length = len(str(columna))
            for fila in datos[:100]:  # Solo revisar las primeras 100 filas para performance
                cell_value = str(fila.get(columna, ''))
                max_length = max(max_length, len(cell_value))
            
            # Limitar el ancho máximo
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[self._get_column_letter(col_idx)].width = adjusted_width
    
    def _limpiar_nombre_hoja(self, nombre: str) -> str:
        """Limpia el nombre para que sea válido como nombre de hoja de Excel"""
        # Caracteres no permitidos en nombres de hojas
        caracteres_invalidos = ['\\', '/', '*', '[', ']', ':', '?']
        nombre_limpio = nombre
        
        for char in caracteres_invalidos:
            nombre_limpio = nombre_limpio.replace(char, '_')
        
        # Eliminar espacios para evitar problemas
        nombre_limpio = nombre_limpio.replace(' ', '_')
        
        # Limitar longitud (máximo 31 caracteres)
        if len(nombre_limpio) > 31:
            nombre_limpio = nombre_limpio[:28] + "..."
        
        return nombre_limpio
    
    def _limpiar_nombre_tabla(self, nombre: str) -> str:
        """Limpia el nombre para que sea válido como nombre de tabla de Excel"""
        # Los nombres de tabla no pueden tener espacios ni caracteres especiales
        caracteres_invalidos = ['\\', '/', '*', '[', ']', ':', '?', ' ', '-', '.', '(', ')']
        nombre_limpio = nombre
        
        for char in caracteres_invalidos:
            nombre_limpio = nombre_limpio.replace(char, '_')
        
        # Debe empezar con letra o guión bajo
        if nombre_limpio and not (nombre_limpio[0].isalpha() or nombre_limpio[0] == '_'):
            nombre_limpio = f"T_{nombre_limpio}"
        
        # Limitar longitud
        if len(nombre_limpio) > 255:
            nombre_limpio = nombre_limpio[:252] + "..."
        
        # Asegurar que no esté vacío
        if not nombre_limpio:
            nombre_limpio = "Tabla_1"
        
        return nombre_limpio
    
    def _limpiar_nombre_archivo(self, nombre: str) -> str:
        """Limpia el nombre para que sea válido como nombre de archivo"""
        # Caracteres no permitidos en nombres de archivo de Windows
        caracteres_invalidos = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        nombre_limpio = nombre
        
        for char in caracteres_invalidos:
            nombre_limpio = nombre_limpio.replace(char, '_')
        
        # Reemplazar espacios por guiones bajos
        nombre_limpio = nombre_limpio.replace(' ', '_')
        
        # Limitar longitud
        if len(nombre_limpio) > 200:
            nombre_limpio = nombre_limpio[:200]
        
        return nombre_limpio
    
    def _get_column_letter(self, col_num: int) -> str:
        """Convierte número de columna a letra de Excel"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result